import os
import tempfile
from typing import Dict, Any, List, Optional, Tuple
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.models.business_model import BusinessModel, BusinessModelField
from app.models.business_model_link import BusinessModelLink
from app.models.data_source import DataSource
from app.dao.action_dao import get_action_dao
from app.utils.db_client import DBClient
from app.utils.data_source_manager import data_source_manager
from app.config import settings
from app.utils.logger import get_logger

logger = get_logger(__name__)

class ExcelImportExportService:
    def __init__(self):
        self.action_dao = get_action_dao()
    
    def export_to_excel_file(self, db: Session) -> Dict[str, Any]:
        """导出业务模型和实例数据到doc目录下的Excel文件"""
        try:
            # 确保doc目录存在
            doc_dir = os.path.join(os.getcwd(), "docs")
            os.makedirs(doc_dir, exist_ok=True)
            
            # 定义输出文件路径
            output_file = os.path.join(doc_dir, "business_model_export.xlsx")
            
            # 创建工作簿
            wb = Workbook()
            
            # 1. 导出对象页签
            self._export_business_models(wb, db)
            
            # 2. 导出对象字段页签
            self._export_business_model_fields(wb, db)
            
            # 3. 导出关系页签
            self._export_business_model_links(wb, db)
            
            # 4. 导出行动页签
            self._export_actions(wb)
            
            # 5. 导出实例数据页签
            self._export_instance_data(wb, db)
            
            # 保存到doc目录
            wb.save(output_file)
            
            return {
                "success": True,
                "file_path": output_file
            }
            
        except Exception as e:
            logger.error(f"Error exporting to Excel file: {e}")
            return {
                "success": False,
                "error": str(e),
                "file_path": None
            }
    
    def import_from_excel_file(self, db: Session) -> Dict[str, Any]:
        """从doc目录下的固定Excel文件导入业务模型和实例数据"""
        try:
            # 定义输入文件路径
            input_file = os.path.join(os.getcwd(), "docs", "business_model_export.xlsx")
            
            if not os.path.exists(input_file):
                return {
                    "success": False,
                    "message": f"Import file not found: {input_file}",
                    "details": {}
                }
            
            # 执行导入
            result = self.import_from_excel(input_file, db)
            return result
            
        except Exception as e:
            logger.error(f"Error importing from Excel file: {e}")
            return {
                "success": False,
                "message": f"Import failed: {str(e)}",
                "details": {}
            }
    
    def export_to_excel(self, db: Session) -> str:
        """导出业务模型和实例数据到Excel文件（临时文件，用于向后兼容）"""
        try:
            # 创建工作簿
            wb = Workbook()
            
            # 1. 导出对象页签
            self._export_business_models(wb, db)
            
            # 2. 导出关系页签
            self._export_business_model_links(wb, db)
            
            # 3. 导出行动页签
            self._export_actions(wb)
            
            # 4. 导出实例数据页签
            self._export_instance_data(wb, db)
            
            # 保存到临时文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_file.name)
            temp_file.close()
            
            return temp_file.name
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def _export_business_models(self, wb: Workbook, db: Session):
        """导出对象页签"""
        ws = wb.active
        ws.title = "对象"
        
        # 表头
        headers = ["ID", "API名称", "中文名称", "中文说明", "主键ID", "数据源ID"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 数据
        business_models = db.query(BusinessModel).all()
        for row, model in enumerate(business_models, 2):
            ws.cell(row=row, column=1, value=model.id)
            ws.cell(row=row, column=2, value=model.api_name)
            ws.cell(row=row, column=3, value=model.name)
            ws.cell(row=row, column=4, value=model.description)
            ws.cell(row=row, column=5, value=model.primary_key_id)
            ws.cell(row=row, column=6, value=model.data_source_id)
        
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
    
    def _export_business_model_fields(self, wb: Workbook, db: Session):
        """导出对象字段页签"""
        ws = wb.create_sheet("对象字段")
            
        # 表头 - 新增"是否必填"、"是否为枚举"、"枚举值"列
        headers = ["ID", "模型ID", "字段ID", "数据类型", "中文名称", "中文说明", "是否必填", "是否为枚举", "枚举值"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        # 数据
        business_model_fields = db.query(BusinessModelField).all()
        for row, field in enumerate(business_model_fields, 2):
            ws.cell(row=row, column=1, value=field.id)
            ws.cell(row=row, column=2, value=field.model_id)
            ws.cell(row=row, column=3, value=field.field_id)
            ws.cell(row=row, column=4, value=field.data_type)
            ws.cell(row=row, column=5, value=field.name)
            ws.cell(row=row, column=6, value=field.description)
            ws.cell(row=row, column=7, value=field.required if field.required is not None else True)
            ws.cell(row=row, column=8, value=field.is_enum if field.is_enum is not None else False)  # 新增：导出is_enum字段
            # 新增：导出enum_values字段（JSON数组转字符串）
            enum_values_str = ''
            if field.enum_values:
                if isinstance(field.enum_values, list):
                    enum_values_str = ', '.join(field.enum_values)
                else:
                    enum_values_str = str(field.enum_values)
            ws.cell(row=row, column=9, value=enum_values_str)
            
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
    
    def _export_business_model_links(self, wb: Workbook, db: Session):
        """导出关系页签"""
        ws = wb.create_sheet("关系")
        
        # 表头
        headers = [
            "ID", "中文名称", "中文说明", "源模型", "源API名称", "源键", 
            "目标模型", "目标API名称", "目标键", "基数", 
            "中间模型", "中间源键", "中间目标键"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 数据
        links = db.query(BusinessModelLink).all()
        for row, link in enumerate(links, 2):
            ws.cell(row=row, column=1, value=link.id)
            ws.cell(row=row, column=2, value=link.name)
            ws.cell(row=row, column=3, value=link.description)
            ws.cell(row=row, column=4, value=link.source_model)
            ws.cell(row=row, column=5, value=link.source_api_name)
            ws.cell(row=row, column=6, value=link.source_key)
            ws.cell(row=row, column=7, value=link.target_model)
            ws.cell(row=row, column=8, value=link.target_api_name)
            ws.cell(row=row, column=9, value=link.target_key)
            ws.cell(row=row, column=10, value=link.cardinality)
            ws.cell(row=row, column=11, value=link.intermediate_model)
            ws.cell(row=row, column=12, value=link.intermediate_source_key)
            ws.cell(row=row, column=13, value=link.intermediate_target_key)
        
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
    
    def _export_actions(self, wb: Workbook):
        """导出行动页签"""
        ws = wb.create_sheet("行动")
        
        # 表头
        headers = [
            "ID", "API名称", "名称", "描述", "动作类型", "操作", "目标模型ID", 
            "目标链接ID", "函数代码", "参数", "提交条件", "创建时间", "更新时间"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 数据
        actions = self.action_dao.get_actions()
        for row, action in enumerate(actions, 2):
            ws.cell(row=row, column=1, value=action.get("id"))
            ws.cell(row=row, column=2, value=action.get("api_name"))
            ws.cell(row=row, column=3, value=action.get("name"))
            ws.cell(row=row, column=4, value=action.get("description"))
            ws.cell(row=row, column=5, value=action.get("action_type"))
            ws.cell(row=row, column=6, value=action.get("operation"))
            ws.cell(row=row, column=7, value=action.get("target_model_id"))
            ws.cell(row=row, column=8, value=action.get("target_link_id"))
            ws.cell(row=row, column=9, value=action.get("function_code"))
            ws.cell(row=row, column=10, value=str(action.get("parameters", [])))
            ws.cell(row=row, column=11, value=str(action.get("submission_criteria", [])))
            ws.cell(row=row, column=12, value=action.get("created_at"))
            ws.cell(row=row, column=13, value=action.get("updated_at"))
        
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
    
    def _export_instance_data(self, wb: Workbook, db: Session):
        """导出实例数据页签"""
        business_models = db.query(BusinessModel).all()
        
        for model in business_models:
            if not model.data_source_id:
                continue
            
            # 获取数据源
            data_source = db.query(DataSource).filter(DataSource.id == model.data_source_id).first()
            if not data_source:
                continue
            
            # 获取表字段信息
            client = DBClient(data_source.type, data_source.connection_string)
            client.connect()
            
            try:
                columns = client.get_table_columns(model.id)
                if not columns:
                    continue
                
                # 创建工作表
                sheet_name = self._sanitize_sheet_name(model.name)
                ws = wb.create_sheet(sheet_name)
                
                # 获取字段映射（英文字段名 -> 中文说明）
                field_mappings = {}
                model_fields = db.query(BusinessModelField).filter(
                    BusinessModelField.model_id == model.id
                ).all()
                for field in model_fields:
                    field_mappings[field.field_id] = field.name
                
                # 双行表头：第一行为英文字段名，第二行为中文说明
                english_headers = []
                chinese_headers = []
                for col_info in columns:
                    field_name = col_info['name']
                    english_headers.append(field_name)
                    chinese_headers.append(field_mappings.get(field_name, field_name))
                
                # 写入第一行（英文字段名）
                for col, header in enumerate(english_headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                
                # 写入第二行（中文说明）
                for col, header in enumerate(chinese_headers, 1):
                    cell = ws.cell(row=2, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                
                # 查询数据
                query = f"SELECT * FROM {model.id}"
                data = client.execute_query(query)
                
                # 写入数据
                for row_idx, record in enumerate(data, 3):
                    for col_idx, field_name in enumerate(english_headers, 1):
                        value = record.get(field_name)
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 自动调整列宽
                for col in range(1, len(english_headers) + 1):
                    ws.column_dimensions[get_column_letter(col)].auto_size = True
                    
            finally:
                client.close()
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """清理工作表名称，确保符合Excel要求"""
        # Excel工作表名称不能超过31个字符，不能包含特殊字符
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        # 截断到31个字符
        if len(name) > 31:
            name = name[:31]
        
        # 如果名称为空，使用默认名称
        if not name.strip():
            name = "Sheet"
        
        return name
    
    def import_from_excel(self, file_path: str, db: Session) -> Dict[str, Any]:
        """从Excel文件导入业务模型和实例数据"""
        try:
            logger.info(f"开始从Excel文件导入数据: {file_path}")
            wb = load_workbook(file_path)
            result = {
                "success": True,
                "message": "Import completed successfully",
                "details": {
                    "business_models": {"imported": 0, "failed": 0, "errors": []},
                    "business_model_fields": {"imported": 0, "failed": 0, "errors": []},
                    "business_model_links": {"imported": 0, "failed": 0, "errors": []},
                    "actions": {"imported": 0, "failed": 0, "errors": []},
                    "instance_data": {"imported": 0, "failed": 0, "errors": []}
                }
            }
            
            # 1. 导入对象配置
            if "对象" in wb.sheetnames:
                logger.info("正在导入对象配置...")
                self._import_business_models(wb["对象"], db, result["details"]["business_models"])
            
            # 2. 导入对象字段配置
            if "对象字段" in wb.sheetnames:
                logger.info("正在导入对象字段配置...")
                self._import_business_model_fields(wb["对象字段"], db, result["details"]["business_model_fields"])
            
            # 3. 导入关系配置
            if "关系" in wb.sheetnames:
                logger.info("正在导入关系配置...")
                self._import_business_model_links(wb["关系"], db, result["details"]["business_model_links"])
            
            # 4. 导入行动配置
            if "行动" in wb.sheetnames:
                logger.info("正在导入行动配置...")
                self._import_actions(wb["行动"], result["details"]["actions"])
            
            # 4. 同步所有业务模型的表结构
            if "对象" in wb.sheetnames and "对象字段" in wb.sheetnames:
                logger.info("正在同步业务模型表结构...")
                self._sync_all_table_structures(wb["对象"], wb["对象字段"], db)
            
            # 5. 导入实例数据
            logger.info("正在导入实例数据...")
            self._import_instance_data(wb, db, result["details"]["instance_data"])
            
            logger.info("Excel文件导入完成")
            return result
            
        except Exception as e:
            logger.error(f"Excel文件导入失败: {e}")
            raise
    
    def _import_business_models(self, ws, db: Session, result_detail: Dict):
        """导入对象配置"""
        try:
            # 获取表头
            headers = [cell.value for cell in ws[1]]
            if not headers or not any(headers):
                return
            
            # 查找必要的列索引
            id_col = self._find_column_index(headers, "ID")
            name_col = self._find_column_index(headers, "中文名称")
            api_name_col = self._find_column_index(headers, "API名称")
            description_col = self._find_column_index(headers, "中文说明")
            primary_key_col = self._find_column_index(headers, "主键ID")
            data_source_col = self._find_column_index(headers, "数据源ID")
            
            if id_col is None or name_col is None:
                result_detail["errors"].append("Missing required columns in business models sheet")
                return
            
            # 处理数据行
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[id_col] or not row[name_col]:
                    continue
                
                try:
                    # 检查是否已存在
                    existing = db.query(BusinessModel).filter(BusinessModel.id == row[id_col]).first()
                    
                    if existing:
                        # 更新现有记录
                        existing.name = row[name_col]
                        existing.api_name = row[api_name_col] if api_name_col is not None else existing.api_name
                        existing.description = row[description_col] if description_col is not None else existing.description
                        existing.primary_key_id = row[primary_key_col] if primary_key_col is not None else existing.primary_key_id
                        existing.data_source_id = row[data_source_col] if data_source_col is not None else existing.data_source_id
                    else:
                        # 创建新记录
                        new_model = BusinessModel(
                            id=row[id_col],
                            name=row[name_col],
                            api_name=row[api_name_col] if api_name_col is not None else None,
                            description=row[description_col] if description_col is not None else None,
                            primary_key_id=row[primary_key_col] if primary_key_col is not None else None,
                            data_source_id=row[data_source_col] if data_source_col is not None else None
                        )
                        db.add(new_model)
                    
                    db.commit()
                    result_detail["imported"] += 1
                    
                except Exception as e:
                    db.rollback()
                    error_msg = f"Failed to import business model {row[id_col]}: {str(e)}"
                    result_detail["errors"].append(error_msg)
                    result_detail["failed"] += 1
            
        except Exception as e:
            error_msg = f"Error processing business models sheet: {str(e)}"
            result_detail["errors"].append(error_msg)
            result_detail["failed"] += 1
    
    def _import_business_model_fields(self, ws, db: Session, result_detail: Dict):
        """导入对象字段配置"""
        try:
            # 获取表头
            headers = [cell.value for cell in ws[1]]
            if not headers or not any(headers):
                return
            
            # 查找必要的列索引
            id_col = self._find_column_index(headers, "ID")
            model_id_col = self._find_column_index(headers, "模型ID")
            field_id_col = self._find_column_index(headers, "字段ID")
            data_type_col = self._find_column_index(headers, "数据类型")
            name_col = self._find_column_index(headers, "中文名称")
            description_col = self._find_column_index(headers, "中文说明")
            required_col = self._find_column_index(headers, "是否必填")
            is_enum_col = self._find_column_index(headers, "是否为枚举")  # 新增：是否为枚举列
            enum_values_col = self._find_column_index(headers, "枚举值")  # 新增：枚举值列
            
            if model_id_col is None or field_id_col is None or name_col is None:
                result_detail["errors"].append("Missing required columns in business model fields sheet")
                return
            
            # 处理数据行
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[model_id_col] or not row[field_id_col] or not row[name_col]:
                    continue
                
                try:
                    # 检查是否已存在（基于ID或模型ID+字段ID）
                    existing = None
                    if id_col is not None and row[id_col]:
                        existing = db.query(BusinessModelField).filter(BusinessModelField.id == row[id_col]).first()
                    
                    if not existing:
                        # 基于模型ID和字段ID查找
                        existing = db.query(BusinessModelField).filter(
                            BusinessModelField.model_id == row[model_id_col],
                            BusinessModelField.field_id == row[field_id_col]
                        ).first()
                    
                    # 处理枚举值：将逗号分隔的字符串转为列表
                    enum_values_list = None
                    if enum_values_col is not None and row[enum_values_col]:
                        enum_values_str = str(row[enum_values_col]).strip()
                        if enum_values_str:
                            enum_values_list = [v.strip() for v in enum_values_str.split(',') if v.strip()]
                    
                    if existing:
                        # 更新现有记录
                        existing.model_id = row[model_id_col]
                        existing.field_id = row[field_id_col]
                        existing.data_type = row[data_type_col] if data_type_col is not None else existing.data_type
                        existing.name = row[name_col]
                        existing.description = row[description_col] if description_col is not None else existing.description
                        # 更新required字段
                        if required_col is not None and row[required_col] is not None:
                            existing.required = bool(row[required_col])
                        # 新增：更新is_enum字段
                        if is_enum_col is not None and row[is_enum_col] is not None:
                            existing.is_enum = bool(row[is_enum_col])
                        # 新增：更新enum_values字段
                        if enum_values_col is not None and enum_values_list is not None:
                            existing.enum_values = enum_values_list if existing.is_enum else None
                    else:
                        # 创建新记录
                        is_enum = bool(row[is_enum_col]) if is_enum_col is not None and row[is_enum_col] is not None else False
                        new_field = BusinessModelField(
                            model_id=row[model_id_col],
                            field_id=row[field_id_col],
                            data_type=row[data_type_col] if data_type_col is not None else "TEXT",
                            name=row[name_col],
                            description=row[description_col] if description_col is not None else None,
                            required=row[required_col] if required_col is not None and row[required_col] is not None else False,
                            is_enum=is_enum,
                            enum_values=enum_values_list if is_enum else None  # 新增：导入enum_values字段
                        )
                        db.add(new_field)
                    
                    db.commit()
                    result_detail["imported"] += 1
                    
                except Exception as e:
                    db.rollback()
                    error_msg = f"Failed to import business model field {row[field_id_col]} for model {row[model_id_col]}: {str(e)}"
                    result_detail["errors"].append(error_msg)
                    result_detail["failed"] += 1
            
        except Exception as e:
            error_msg = f"Error processing business model fields sheet: {str(e)}"
            result_detail["errors"].append(error_msg)
            result_detail["failed"] += 1
    
    def _import_business_model_links(self, ws, db: Session, result_detail: Dict):
        """导入关系配置"""
        try:
            # 获取表头
            headers = [cell.value for cell in ws[1]]
            if not headers or not any(headers):
                return
            
            # 查找必要的列索引
            id_col = self._find_column_index(headers, "ID")
            name_col = self._find_column_index(headers, "中文名称")
            source_model_col = self._find_column_index(headers, "源模型")
            target_model_col = self._find_column_index(headers, "目标模型")
            cardinality_col = self._find_column_index(headers, "基数")
            
            if id_col is None or name_col is None or source_model_col is None or target_model_col is None or cardinality_col is None:
                result_detail["errors"].append("Missing required columns in business model links sheet")
                return
            
            # 处理数据行
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[id_col] or not row[name_col]:
                    continue
                
                try:
                    # 检查是否已存在
                    existing = db.query(BusinessModelLink).filter(BusinessModelLink.id == row[id_col]).first()
                    
                    if existing:
                        # 更新现有记录
                        existing.name = row[name_col]
                        existing.source_model = row[source_model_col]
                        existing.target_model = row[target_model_col]
                        existing.cardinality = row[cardinality_col]
                        # 更新其他可选字段
                        self._update_optional_link_fields(existing, row, headers)
                    else:
                        # 创建新记录
                        new_link = BusinessModelLink(
                            id=row[id_col],
                            name=row[name_col],
                            source_model=row[source_model_col],
                            target_model=row[target_model_col],
                            cardinality=row[cardinality_col]
                        )
                        # 设置其他可选字段
                        self._update_optional_link_fields(new_link, row, headers)
                        db.add(new_link)
                    
                    db.commit()
                    result_detail["imported"] += 1
                    
                except Exception as e:
                    db.rollback()
                    error_msg = f"Failed to import business model link {row[id_col]}: {str(e)}"
                    result_detail["errors"].append(error_msg)
                    result_detail["failed"] += 1
            
        except Exception as e:
            error_msg = f"Error processing business model links sheet: {str(e)}"
            result_detail["errors"].append(error_msg)
            result_detail["failed"] += 1
    
    def _update_optional_link_fields(self, link: BusinessModelLink, row: tuple, headers: List[str]):
        """更新关系的可选字段"""
        source_api_col = self._find_column_index(headers, "源API名称")
        source_key_col = self._find_column_index(headers, "源键")
        target_api_col = self._find_column_index(headers, "目标API名称")
        target_key_col = self._find_column_index(headers, "目标键")
        intermediate_model_col = self._find_column_index(headers, "中间模型")
        intermediate_source_col = self._find_column_index(headers, "中间源键")
        intermediate_target_col = self._find_column_index(headers, "中间目标键")
        description_col = self._find_column_index(headers, "中文说明")
        
        if source_api_col is not None:
            link.source_api_name = row[source_api_col]
        if source_key_col is not None:
            link.source_key = row[source_key_col]
        if target_api_col is not None:
            link.target_api_name = row[target_api_col]
        if target_key_col is not None:
            link.target_key = row[target_key_col]
        if intermediate_model_col is not None:
            link.intermediate_model = row[intermediate_model_col]
        if intermediate_source_col is not None:
            link.intermediate_source_key = row[intermediate_source_col]
        if intermediate_target_col is not None:
            link.intermediate_target_key = row[intermediate_target_col]
        if description_col is not None:
            link.description = row[description_col]
    
    def _import_actions(self, ws, result_detail: Dict):
        """导入行动配置"""
        try:
            # 获取表头
            headers = [cell.value for cell in ws[1]]
            if not headers or not any(headers):
                return
            
            # 中文表头到英文字段名的映射
            header_to_field_map = {
                "ID": "id",
                "API名称": "api_name",
                "名称": "name",
                "描述": "description",
                "动作类型": "action_type",
                "操作": "operation",
                "目标模型ID": "target_model_id",
                "目标链接ID": "target_link_id",
                "函数代码": "function_code",
                "参数": "parameters",
                "提交条件": "submission_criteria",
                "创建时间": "created_at",
                "更新时间": "updated_at"
            }
            
            # 查找必要的列索引
            id_col = self._find_column_index(headers, "ID")
            name_col = self._find_column_index(headers, "名称")
            
            if id_col is None or name_col is None:
                result_detail["errors"].append("Missing required columns in actions sheet")
                return
            
            # 处理数据行
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[id_col] or not row[name_col]:
                    continue
                
                try:
                    action_data = {}
                    for i, header in enumerate(headers):
                        if header and i < len(row):
                            value = row[i]
                            # 获取对应的英文字段名
                            field_name = header_to_field_map.get(header, header)
                            
                            if header == "参数" and isinstance(value, str):
                                # 尝试解析参数字符串
                                try:
                                    import ast
                                    action_data[field_name] = ast.literal_eval(value)
                                except:
                                    action_data[field_name] = []
                            elif header == "提交条件" and isinstance(value, str):
                                # 尝试解析提交条件字符串
                                try:
                                    import ast
                                    action_data[field_name] = ast.literal_eval(value)
                                except:
                                    action_data[field_name] = []
                            elif header in ["创建时间", "更新时间"]:
                                # 跳过时间字段，让DAO自动处理
                                continue
                            else:
                                action_data[field_name] = value
                    
                    # 确保必要字段存在
                    action_data["id"] = row[id_col]
                    action_data["name"] = row[name_col]
                    
                    # 检查是否已存在
                    existing = self.action_dao.get_action_by_id(row[id_col])
                    if existing:
                        # 更新现有记录
                        success = self.action_dao.update_action(row[id_col], action_data)
                    else:
                        # 创建新记录
                        success = self.action_dao.create_action(action_data) is not None
                    
                    if success:
                        result_detail["imported"] += 1
                    else:
                        result_detail["failed"] += 1
                        result_detail["errors"].append(f"Failed to import action {row[id_col]}")
                        
                except Exception as e:
                    error_msg = f"Failed to import action {row[id_col]}: {str(e)}"
                    result_detail["errors"].append(error_msg)
                    result_detail["failed"] += 1
            
        except Exception as e:
            error_msg = f"Error processing actions sheet: {str(e)}"
            result_detail["errors"].append(error_msg)
            result_detail["failed"] += 1
    
    def _import_instance_data(self, wb, db: Session, result_detail: Dict):
        """导入实例数据"""
        try:
            # 获取所有业务模型
            business_models = {model.id: model for model in db.query(BusinessModel).all()}
            
            # 遍历所有工作表（除了标准页签）
            standard_sheets = {"对象", "关系", "行动"}
            for sheet_name in wb.sheetnames:
                logger.info(f"正在导入工作表: {sheet_name}")
                if sheet_name in standard_sheets:
                    continue
                
                # 找到对应的业务模型（通过中文名称匹配）
                model = None
                for bm in business_models.values():
                    if bm.name == sheet_name:
                        model = bm
                        break
                
                if not model or not model.data_source_id:
                    continue
                
                ws = wb[sheet_name]
                self._import_instance_data_for_model(ws, model, db, result_detail)
                
        except Exception as e:
            error_msg = f"Error processing instance data sheets: {str(e)}"
            result_detail["errors"].append(error_msg)
            result_detail["failed"] += 1
    
    def _import_instance_data_for_model(self, ws, model: BusinessModel, db: Session, result_detail: Dict):
        """为特定模型导入实例数据"""
        try:
            # 获取双行表头
            english_headers = [cell.value for cell in ws[1]]  # 第一行：英文字段名
            if not english_headers or not any(english_headers):
                return
            
            # 验证表头
            valid_headers = []
            for header in english_headers:
                if header and isinstance(header, str):
                    valid_headers.append(header)
                else:
                    valid_headers.append(None)
            
            # 获取数据源
            data_source = db.query(DataSource).filter(DataSource.id == model.data_source_id).first()
            if not data_source:
                result_detail["errors"].append(f"Data source not found for model {model.id}")
                return
            
            # 收集所有数据记录
            records_to_import = []
            row_indices = []  # 记录对应的行号，用于错误报告
            
            # 处理数据行（从第3行开始）
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
                try:
                    # 构建数据记录
                    record = {}
                    for col_idx, field_name in enumerate(valid_headers):
                        if field_name and col_idx < len(row):
                            record[field_name] = row[col_idx]
                    
                    if not record:
                        continue
                    
                    records_to_import.append(record)
                    row_indices.append(row_idx)
                    
                except Exception as e:
                    result_detail["failed"] += 1
                    result_detail["errors"].append(f"Error preparing record at row {row_idx} for model {model.id}: {str(e)}")
            
            if not records_to_import:
                return
            
            logger.info(f"批量导入 {len(records_to_import)} 条记录到模型 {model.id}")
            
            # 批量执行UPSERT操作
            primary_key_field = model.primary_key_id
            batch_result = data_source_manager.execute_batch_upsert(
                data_source_id=model.data_source_id,
                table_name=model.id,
                data_list=records_to_import,
                primary_key=primary_key_field
            )
            
            result_detail["imported"] += batch_result["success_count"]
            result_detail["failed"] += batch_result["failed_count"]
            
            if batch_result["failed_count"] > 0:
                # 简单记录失败信息（由于批量操作，无法精确定位到具体行）
                result_detail["errors"].append(f"批量导入失败 {batch_result['failed_count']} 条记录到模型 {model.id}")
            
        except Exception as e:
            error_msg = f"Error importing instance data for model {model.id}: {str(e)}"
            result_detail["errors"].append(error_msg)
            result_detail["failed"] += len(records_to_import) if 'records_to_import' in locals() else 1
    
    def _sync_all_table_structures(self, business_models_ws, business_model_fields_ws, db: Session):
        """同步所有业务模型的表结构"""
        try:
            # 获取业务模型字段数据
            field_headers = [cell.value for cell in business_model_fields_ws[1]]
            model_id_col = self._find_column_index(field_headers, "模型ID")
            field_id_col = self._find_column_index(field_headers, "字段ID")
            data_type_col = self._find_column_index(field_headers, "数据类型")
            name_col = self._find_column_index(field_headers, "中文名称")
            description_col = self._find_column_index(field_headers, "中文说明")
            
            if model_id_col is None or field_id_col is None:
                logger.warning("对象字段页签缺少必要列，跳过表结构同步")
                return
            
            # 构建模型字段映射
            model_fields_map = {}
            for row in business_model_fields_ws.iter_rows(min_row=2, values_only=True):
                if not row[model_id_col] or not row[field_id_col]:
                    continue
                
                model_id = row[model_id_col]
                if model_id not in model_fields_map:
                    model_fields_map[model_id] = []
                
                field_dict = {
                    'field_id': row[field_id_col],
                    'data_type': row[data_type_col] if data_type_col is not None else "TEXT",
                    'name': row[name_col] if name_col is not None else row[field_id_col],
                    'description': row[description_col] if description_col is not None else None
                }
                model_fields_map[model_id].append(field_dict)
            
            # 获取业务模型数据
            model_headers = [cell.value for cell in business_models_ws[1]]
            id_col = self._find_column_index(model_headers, "ID")
            primary_key_col = self._find_column_index(model_headers, "主键ID")
            data_source_col = self._find_column_index(model_headers, "数据源ID")
            
            if id_col is None:
                logger.warning("对象页签缺少ID列，跳过表结构同步")
                return
            
            # 同步每个业务模型的表结构
            for row in business_models_ws.iter_rows(min_row=2, values_only=True):
                if not row[id_col]:
                    continue
                
                model_id = row[id_col]
                data_source_id = row[data_source_col] if data_source_col is not None else None
                primary_key_id = row[primary_key_col] if primary_key_col is not None else None
                
                if not data_source_id:
                    continue
                
                model_fields = model_fields_map.get(model_id, [])
                
                success = data_source_manager.sync_table_structure(
                    data_source_id=data_source_id,
                    table_name=model_id,
                    model_fields=model_fields,
                    primary_key=primary_key_id
                )
                
                if not success:
                    logger.warning(f"同步表结构失败: {model_id}")
                    
        except Exception as e:
            logger.error(f"同步所有表结构失败: {str(e)}")
    
    def _find_column_index(self, headers: List[str], target_header: str) -> Optional[int]:
        """在表头列表中查找目标列的索引"""
        for i, header in enumerate(headers):
            if header == target_header:
                return i
        return None

def get_excel_service() -> ExcelImportExportService:
    return ExcelImportExportService()